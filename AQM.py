import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import constants
sns.set(rc={"figure.dpi":300, 'savefig.dpi':300})
sns.set_context('notebook')
sns.set_style("ticks")

h = constants.h
k = constants.k
c = constants.c

T = 6040
l = np.linspace(0.001, 2500, 2501)

#Planck's law (in terms of wavelength)

B = ((2*h*c**2)/(l*(l*10**(-9))**4))*(1/(np.exp((h*c)/((l*10**(-9))*k*T)) - 1))*6.8*10**(-5)

T2 = 5780
B2 = ((2*h*c**2)/(l*(l*10**(-9))**4))*(1/(np.exp((h*c)/((l*10**(-9))*k*T2)) - 1))*6.8*10**(-5)


import openpyxl
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("Irradiancedata.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active

data = []
lamda1 = np.zeros(4210)
irradiation1 = np.zeros(4210)
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        #print(col[row].value)
        data.append(col[row].value)

for i in range(2, 4210):
    if i%2 == 0:
        lamda1[i] = data[i]
    else:
        irradiation1[i] = data[i]

lamda2 = lamda1[::2] 
irradiation2 = irradiation1[1::2]

lamda = np.zeros(2104)
irradiation = np.zeros(2104)

for i in range (0, 2104):
    lamda[i] = lamda2[i+1]
    irradiation[i] = irradiation2[i+1]

# Plotting
plt.plot(l, B, color='brown', linestyle = '-.', label=r'Planck´s Law ($T$=6040K)')
plt.plot(l, B2, color='black', linestyle = '--', label=r'Planck´s Law ($T$=5780K)')
plt.plot(lamda, irradiation, 'b', lw = '1', label=r'Solar irradiance data')
plt.legend(loc="upper right")
plt.xlabel(r'$\lambda$ (nm)')
plt.ylabel(r'Spectral irradiance (W/m$^2$/nm)')
plt.savefig('SolarIrradiation.pdf') 
plt.show()